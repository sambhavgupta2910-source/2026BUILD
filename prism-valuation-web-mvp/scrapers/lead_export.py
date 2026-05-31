"""
PRISM Lead Engine — Export to Excel + PRISM API sync

Merges leads from:
  1. leads_sellers.json  (seller_dubizzle.py output)
  2. leads_buyers.json   (buyer_reddit.py output)
  3. leads.csv           (PRISM in-app inquiries)

Produces Dubai_Leads.xlsx with:
  - Dashboard tab  (summary stats + grade breakdown)
  - Sellers tab    (scored seller leads)
  - Buyers tab     (scored buyer leads)
  - Inquiries tab  (PRISM in-app form submissions)

Optionally pushes in-app inquiry leads to PRISM server via /api/inquiry.

Usage:
    pip install openpyxl requests
    python lead_export.py [--prism http://localhost:4173]
"""

import argparse
import csv
import json
import os
from datetime import datetime, timezone

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[warn] openpyxl not installed — Excel export disabled. Run: pip install openpyxl")

GRADE_COLORS = {"A": "FF4ECBA8", "B": "FF7DB5FF", "C": "FFF0B440", "D": "FF8FA3C8"}
DARK_BG = "FF0F1623"
HEADER_BG = "FF182030"
TEXT_WHITE = "FFDDE6F5"
ACCENT = "FF5B9CF6"


def load_json(path):
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def load_csv_inquiries(path):
    if not os.path.exists(path):
        return []
    rows = []
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(dict(row))
    return rows


def _style_header_row(ws, row_num, col_count):
    fill = PatternFill("solid", fgColor=HEADER_BG)
    font = Font(bold=True, color=TEXT_WHITE, size=10)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _grade_fill(grade):
    color = GRADE_COLORS.get(grade, "FF5A7099")
    return PatternFill("solid", fgColor=color)


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_sellers_tab(wb, sellers):
    ws = wb.create_sheet("Sellers")
    ws.sheet_properties.tabColor = "4ECBA8"
    headers = ["Grade", "Score", "Title", "Price (AED)", "Location", "Days Listed",
               "Phone", "Price Drop", "URL", "Scraped At"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for s in sellers:
        row = [
            s.get("grade", ""),
            s.get("score", 0),
            s.get("title", ""),
            s.get("price_aed", 0),
            s.get("location", ""),
            s.get("days_listed", ""),
            s.get("phone", ""),
            "Yes" if s.get("price_drop") else "No",
            s.get("url", ""),
            s.get("scraped_at", "")[:10],
        ]
        ws.append(row)
        grade = s.get("grade", "D")
        ws.cell(row=ws.max_row, column=1).fill = _grade_fill(grade)
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF090D14")
    _set_col_widths(ws, [8, 7, 50, 14, 22, 12, 16, 10, 50, 12])


def write_buyers_tab(wb, buyers):
    ws = wb.create_sheet("Buyers")
    ws.sheet_properties.tabColor = "5B9CF6"
    headers = ["Grade", "Score", "Title", "Budget (AED)", "Budget Label", "Subreddit",
               "Comments", "Upvotes", "Hours Old", "Author", "URL", "Scraped At"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for b in buyers:
        row = [
            b.get("grade", ""),
            b.get("score", 0),
            b.get("title", ""),
            b.get("budget_aed", 0),
            b.get("budget_label", ""),
            b.get("subreddit", ""),
            b.get("comments", 0),
            b.get("upvotes", 0),
            b.get("hours_old", ""),
            b.get("author", ""),
            b.get("url", ""),
            b.get("scraped_at", "")[:10],
        ]
        ws.append(row)
        grade = b.get("grade", "D")
        ws.cell(row=ws.max_row, column=1).fill = _grade_fill(grade)
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF090D14")
    _set_col_widths(ws, [8, 7, 50, 14, 14, 16, 10, 10, 10, 18, 50, 12])


def write_inquiries_tab(wb, inquiries):
    ws = wb.create_sheet("Inquiries")
    ws.sheet_properties.tabColor = "F0B440"
    headers = ["Timestamp", "Name", "WhatsApp", "Intent", "Area", "Budget", "Message", "Listing ID"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for inq in inquiries:
        row = [
            inq.get("timestamp", ""),
            inq.get("name", ""),
            inq.get("whatsapp", ""),
            inq.get("intent", ""),
            inq.get("area", ""),
            inq.get("budget", ""),
            inq.get("message", ""),
            inq.get("listingId", ""),
        ]
        ws.append(row)
    _set_col_widths(ws, [22, 22, 18, 12, 20, 18, 50, 16])


def write_dashboard_tab(wb, sellers, buyers, inquiries):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_properties.tabColor = ACCENT[2:]

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws["A1"] = "PRISM Lead Engine — Dashboard"
    ws["A1"].font = Font(bold=True, size=16, color=ACCENT[2:])
    ws["A2"] = f"Generated: {now}"
    ws["A2"].font = Font(size=10, color="FF8FA3C8")

    ws["A4"] = "Summary"
    ws["A4"].font = Font(bold=True, size=12, color=TEXT_WHITE)

    stats = [
        ("Total Sellers", len(sellers)),
        ("Total Buyers", len(buyers)),
        ("In-App Inquiries", len(inquiries)),
        ("Grade A Sellers", sum(1 for s in sellers if s.get("grade") == "A")),
        ("Grade A Buyers", sum(1 for b in buyers if b.get("grade") == "A")),
        ("Sellers with Phone", sum(1 for s in sellers if s.get("phone"))),
        ("Buyers > AED 2M", sum(1 for b in buyers if b.get("budget_aed", 0) >= 2_000_000)),
        ("Price Drop Listings", sum(1 for s in sellers if s.get("price_drop"))),
    ]

    ws["A5"] = "Metric"
    ws["B5"] = "Count"
    ws["A5"].font = Font(bold=True, color=TEXT_WHITE)
    ws["B5"].font = Font(bold=True, color=TEXT_WHITE)
    ws["A5"].fill = PatternFill("solid", fgColor=HEADER_BG)
    ws["B5"].fill = PatternFill("solid", fgColor=HEADER_BG)

    for i, (label, val) in enumerate(stats, 6):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = val
        ws[f"A{i}"].font = Font(color=TEXT_WHITE)
        ws[f"B{i}"].font = Font(bold=True, color=TEXT_WHITE)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12


def main():
    parser = argparse.ArgumentParser(description="Export PRISM leads to Excel")
    parser.add_argument("--sellers", default="leads_sellers.json")
    parser.add_argument("--buyers", default="leads_buyers.json")
    parser.add_argument("--inquiries-csv", default="../leads.csv", help="Path to server leads.csv")
    parser.add_argument("--output", default="Dubai_Leads.xlsx")
    args = parser.parse_args()

    sellers = load_json(args.sellers)
    buyers = load_json(args.buyers)
    inquiries = load_csv_inquiries(args.inquiries_csv)

    print(f"[load] {len(sellers)} sellers, {len(buyers)} buyers, {len(inquiries)} in-app inquiries")

    if not HAS_OPENPYXL:
        print("[error] Install openpyxl to generate Excel: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_dashboard_tab(wb, sellers, buyers, inquiries)
    write_sellers_tab(wb, sellers)
    write_buyers_tab(wb, buyers)
    write_inquiries_tab(wb, inquiries)

    wb.save(args.output)
    print(f"[done] {args.output} written — {len(sellers) + len(buyers) + len(inquiries)} total leads")
    print(f"  Tabs: Dashboard · Sellers · Buyers · Inquiries")


if __name__ == "__main__":
    main()
